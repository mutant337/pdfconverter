from openpyxl import load_workbook
from operator import itemgetter
from tkinter import filedialog
from PIL import Image
from PyPDF2 import PdfFileMerger, PdfFileWriter, PdfFileReader
from pagination import add_text_to_pdf
import tkinter as tk
import os
import sys
import comtypes.client
import img2pdf
import pdfkit
import shutil


# Определение кода формата PDF
wdFormatPDF = 17
flag = True
temp = os.path.join(os.path.expanduser("~"), "tempdocs")

if not os.path.exists(temp):
    os.makedirs(temp)

def word_to_pdf(input, output):
    word = comtypes.client.CreateObject('Word.Application')
    doc = word.Documents.Open(input)
    doc.SaveAs(output, FileFormat=wdFormatPDF)
    doc.Close()
    word.Quit()

def excel_to_pdf(input, output):
    # Создание объекта Excel
    excel = comtypes.client.CreateObject('Excel.Application')

    # Открываем книгу Excel
    workbook = excel.Workbooks.Open(input)

    # Сохраняем книгу в формате PDF
    workbook.ExportAsFixedFormat(0, output)

    # Закрываем книгу
    workbook.Close()

    # Закрываем приложение Excel
    excel.Quit()

def img_to_pdf(input, output):
    # opening image
    image = Image.open(input)
    
    # converting into chunks using img2pdf
    pdf_bytes = img2pdf.convert(image.filename)
    
    # opening or creating pdf file
    file = open(output, "wb")
    
    # writing pdf files with chunks
    file.write(pdf_bytes)
    
    # closing image file
    image.close()

    # closing pdf file
    file.close()

def html_to_pdf(input, output):
    path_to_wkhtmltopdf = 'C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)
    pdfkit.from_file(input, output_path=output, configuration=config)

# Загрузка Excel-файла
root = tk.Tk()
root.withdraw()
excel_path = filedialog.askopenfilename(title="Выберіть файл Excel", filetypes=[("Excel Files", "*.xlsx")])

# Если пользователь выбрал файл, то продолжаем работу с ним
if excel_path:
    workbook = load_workbook(excel_path)
    print('Файл вдало завантажен')
else:
    print('Некоректні дані')
    # Здесь можно продолжить обработку файла, как в вашем исходном коде
    # Например, получить список листов: sheets = workbook.sheetnames
# Закрытие окна
root.destroy()

sheet = workbook.active

# Створення словника
data_dict = {}

skip = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    source_path, priority, destination, end = row[0], row[1], row[2], row[4]

    if skip:
        if end:
            skip = 0
        continue

    if source_path:  # Если ячейка непустая
        try:
            # Здесь указываем путь к файлу, который хотим открыть
            with open(source_path, 'rb') as file:
                # Обрабатываем содержимое файла
                # Ваш код, который обрабатывает содержимое файла, если он успешно открыт
                print(f"Файл {source_path} успешно открыт!")
                data_dict[priority] = source_path
        except Exception as e:
            # Общий обработчик ошибок для других возможных исключений
            print(f"Ошибка при открытии файла: {e}")
            print("---!!!---SKIP---!!!---")
            data_dict = {}
            skip = 1
            continue
       
        if destination:
            result = destination
            # Разделение пути на компоненты и получение только имени папки
            result_folder_name = os.path.dirname(result)
            result_file_name = os.path.basename(result)
            if not os.path.exists(result_folder_name): # Перевірка на існування destination
                os.makedirs(result_folder_name)

        if end:
            sorted_dict = {key: value for key, value in sorted(data_dict.items())}
            merger = PdfFileMerger()
            for i in sorted_dict:
                # Получаем расширение файла
                file_path = sorted_dict[i]
                file_extension = os.path.splitext(file_path)[1]

                # Получаем только имя файла без расширения из пути
                file_name = os.path.basename(file_path)
                file_name_without_extension = os.path.splitext(file_name)[0]

                # Добавляем ".pdf" к названию документа без расширения и получаем путь для сохранения временного файла
                converted_file_path = temp + '\\' + file_name + ".pdf"

                print(f'Обробка {file_path}')
                # Проводим конвертацию файла во временную папку в зависимости от расширения
                try:
                    if file_extension == '.pdf':
                        shutil.copy(file_path, converted_file_path)
                
                    elif file_extension in ['.doc', '.docx']:
                        word_to_pdf(file_path, converted_file_path)

                    elif file_extension in ['.xls', '.xlsx']:
                        excel_to_pdf(file_path, converted_file_path)

                    elif file_extension in ['.jpg', '.png', 'jpeg']:
                        img_to_pdf(file_path, converted_file_path)

                    elif file_extension in ['.html']:
                        html_to_pdf(file_path, converted_file_path)

                    merger.append(converted_file_path)

                except Exception as e:
                    print(f"Помилка конвертації {file_path}: {e}")
                    print("---!!!---SKIP---!!!---")
                    data_dict, result = {}, ''
                    flag = False
                    break
            
            if flag:
                merger.write(result)
                merger.close()
                merger = PdfFileMerger()
                # adding blank page
                a = open(result, 'rb')
                pdf=PdfFileReader(a)
                numPages=pdf.getNumPages()
                if numPages % 2 != 0:
                    outPdf=PdfFileWriter()
                    outPdf.appendPagesFromReader(pdf)
                    #outPdf.cloneDocumentFromReader(pdf)
                    outPdf.addBlankPage()
                    outStream=open('Amended.pdf','wb')
                    outPdf.write(outStream)
                    outStream.close()
                    a.close()
                    #Copy amended file back over the original
                    shutil.copyfile('Amended.pdf',result)
                    os.remove("Amended.pdf")
                else:
                    a.close()
                # adding paginatin

                add_text_to_pdf(result, result+'1.pdf')
                os.remove(result)
                os.rename(result+'1.pdf', result)
                
                print(f'---!!!---Результат успішно збережен---!!!--- {result}')
                data_dict, result = {}, ''

            flag = True
                
print('---!!!---Завершення програми---!!!---')


    

    

    

# def compress(image_file):

#     filepath = os.path.join(os.getcwd(), image_file)

#     image = Image.open(filepath)

#     image.save("C:/Users/kostiantyn.dzhelalov/Desktop/image-file-compressed.jpg",
#                  "JPEG",
#                  optimize = True,
#                  quality = 10)
#     return

# compress("C:/Users/kostiantyn.dzhelalov/Desktop/Labels.png")
